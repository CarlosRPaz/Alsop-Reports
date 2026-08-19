"""Inspect the Quotes sheet structure and test dedup logic"""
import openpyxl
from datetime import datetime

DSR_PATH = r"c:\Users\scag3s29\Documents\Claude Scope\Daily Standup Report.xlsx"

wb = openpyxl.load_workbook(DSR_PATH, data_only=True, read_only=False)
ws = wb["Quotes"]

# Get headers
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("COLUMNS:")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

# Show 3 sample rows
print("\nSAMPLE ROWS:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
    print(f"\n  Row {i}:")
    for j, (h, v) in enumerate(zip(headers, row)):
        if v is not None:
            print(f"    [{j}] {h}: {repr(v)}")

# Count total May 2026 rows
def parse_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try: return datetime.strptime(val.strip(), fmt)
            except: pass
    return None

may_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    # Check all date-like columns
    for ci in range(len(headers)):
        v = row[ci] if ci < len(row) else None
        d = parse_date(v)
        if d and d.year == 2026 and d.month == 5:
            may_count += 1
            break

print(f"\nTotal May 2026 quote rows: {may_count}")
print(f"Total rows: {ws.max_row - 1}")

wb.close()
