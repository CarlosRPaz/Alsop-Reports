"""
Fix Rosie's AgencyZoom Name in the Spine: "Rosario Delgado" -> "Rosie Delgado"
Also re-backfill the 4 agents whose premium was zeroed (Estela, Robert, Rosie, Lori)
because the backfill script used exact match instead of Spine resolver.
"""
import os, json, sys
import pandas as pd
import requests
import warnings
import openpyxl
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding='utf-8')

DSR_PATH = r"C:\Users\scag3s29\Documents\Claude Scope\Daily Standup Report.xlsx"

def load_config():
    with open("config/config.json") as f:
        return json.load(f)

config = load_config()
url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or config.get("supabase", {}).get("url")
key = os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY") or config.get("supabase", {}).get("key")
headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates"}

# ---- 1. Fix Rosie's AZ Name in Spine ----
print("[1] Fixing Rosie's AZ Name in Spine sheet...")
wb = openpyxl.load_workbook(DSR_PATH)
ws = wb["Spine"]

# Find the header row and AgencyZoom Name column
az_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "AgencyZoom Name":
        az_col = col
        break

if az_col:
    for row in range(2, ws.max_row + 1):
        agent = ws.cell(row=row, column=1).value
        if agent == "Rosie":
            old_val = ws.cell(row=row, column=az_col).value
            ws.cell(row=row, column=az_col).value = "Rosie Delgado"
            print(f"  Fixed: '{old_val}' -> 'Rosie Delgado'")
            break

wb.save(DSR_PATH)
print("  Saved Spine")

# ---- 2. Re-backfill premium for the 4 agents with name issues ----
print("\n[2] Re-backfilling premium for Estela, Lori, Robert, Rosie...")

# Read Premium data with Spine resolver (using updated Spine)
from src.spine import Spine
spine = Spine(DSR_PATH, sheet_name="Spine", excluded_agents=["Teyssy", "Elizabeth"])

prem_df = pd.read_excel(DSR_PATH, sheet_name="Premium")
prem_df["DateOnly"] = pd.to_datetime(prem_df["Date"]).dt.date

from datetime import date
april_prem = prem_df[(prem_df["DateOnly"] >= date(2026, 4, 1)) & (prem_df["DateOnly"] <= date(2026, 4, 29))]

# Get agent IDs
r = requests.get(f"{url}/rest/v1/agents?select=id,name", headers=headers)
name_to_id = {a["name"]: a["id"] for a in r.json()}

# Resolve each premium entry through Spine
target_agents = {"Estela", "Lori", "Robert", "Rosie"}
updates_needed = {}

for _, row in april_prem.iterrows():
    producer = str(row.get("Producer", "")).strip()
    agent = spine.resolve_agent(producer)
    if agent and agent in target_agents:
        d = row["DateOnly"].isoformat()
        premium = float(row.get("Premium", 0) or 0)
        prem_items = int(row.get("Items", 0) or 0)
        prem_points = float(row.get("Points", 0) or 0)
        key = (agent, d)
        if key not in updates_needed:
            updates_needed[key] = {"premium": 0, "items": 0, "points": 0}
        updates_needed[key]["premium"] += premium
        updates_needed[key]["items"] += prem_items
        updates_needed[key]["points"] += prem_points

print(f"  Found {len(updates_needed)} (agent, date) pairs to update")

for (agent, d), data in sorted(updates_needed.items()):
    agent_id = name_to_id.get(agent)
    if not agent_id:
        print(f"  No agent_id for {agent}")
        continue
    
    r = requests.patch(
        f"{url}/rest/v1/daily_metrics?agent_id=eq.{agent_id}&report_date=eq.{d}",
        headers=headers,
        json={"prem_premium": data["premium"], "prem_items": data["items"], "prem_points": data["points"]}
    )
    status = "OK" if r.status_code < 400 else f"ERR: {r.text[:100]}"
    print(f"  {agent:12s} {d}  prem=${data['premium']:>8,.0f}  items={data['items']}  -> {status}")

# ---- 3. Verify ----
print("\n[3] Verifying totals...")
r = requests.get(
    f"{url}/rest/v1/daily_metrics?select=agent_id,items,prem_premium&report_date=gte.2026-04-01&report_date=lte.2026-04-29",
    headers=headers
)
total_items = sum(m.get("items") or 0 for m in r.json())
total_prem = sum(float(m.get("prem_premium") or 0) for m in r.json())
print(f"  Items MTD: {total_items}")
print(f"  Premium MTD: ${total_prem:,.0f}")
