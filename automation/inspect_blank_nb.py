"""Dump ALL columns for the blank-name NB policies"""
import openpyxl
from datetime import datetime

DSR_PATH = r"c:\Users\scag3s29\Documents\Claude Scope\Daily Standup Report.xlsx"

def parse_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try: return datetime.strptime(val.strip(), fmt)
            except: pass
    return None

wb = openpyxl.load_workbook(DSR_PATH, data_only=True, read_only=False)
ws = wb["NB"]

# Get header row
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("COLUMNS:")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

print("\n" + "=" * 100)
print("BLANK-NAME NB RECORDS (May 2026):")
print("=" * 100)

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name = (row[2] or "").strip()
    d = parse_date(row[7])
    if d and d.year == 2026 and d.month == 5 and not name:
        print(f"\n--- Row {row_idx} ---")
        for i, (header, val) in enumerate(zip(headers, row)):
            if val is not None:
                print(f"  [{i}] {header}: {repr(val)}")

# Also check: are there other blank-name records in recent months?
print("\n\n" + "=" * 100)
print("ALL BLANK-NAME RECORDS (any date):")
print("=" * 100)
blank_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name = (row[2] or "").strip()
    if not name:
        d = parse_date(row[7])
        blank_count += 1
        # Show the Bind ID Name (col 4) and Agent Number (col 0) which might help identify
        agent_num = row[0]
        bind_name = row[4]
        customer = row[6]
        items = row[14]
        date_str = d.strftime("%Y-%m-%d") if d else repr(row[7])
        print(f"  Row {row_idx}: date={date_str}, agent#={agent_num}, bind_id_name={bind_name}, customer={customer}, items={items}")

print(f"\nTotal blank-name records: {blank_count}")

wb.close()
