import openpyxl

path = r"c:\Users\scag3s29\Documents\Claude Scope\Daily Standup Report.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
ws = wb["NB"]

# Check the last 5 rows - show ALL cell values and types
print("Last 10 rows - raw values and types:")
for row in ws.iter_rows(min_row=ws.max_row - 9, max_row=ws.max_row, values_only=False):
    for cell in row:
        if cell.column_letter in ("H", "I"):  # Date columns
            print(f"  Row {cell.row}, Col {cell.column_letter}: value={repr(cell.value)}, type={type(cell.value).__name__}, number_format={cell.number_format}")

# Also check a row from the beginning that DID have dates
print("\nRow 2 date columns for comparison:")
for row in ws.iter_rows(min_row=2, max_row=2, values_only=False):
    for cell in row:
        if cell.column_letter in ("H", "I"):
            print(f"  Row {cell.row}, Col {cell.column_letter}: value={repr(cell.value)}, type={type(cell.value).__name__}, number_format={cell.number_format}")

# Try reading with data_only=False to see formulas
wb2 = openpyxl.load_workbook(path, data_only=False, read_only=False)
ws2 = wb2["NB"]
print("\nLast 5 rows with data_only=False:")
for row in ws2.iter_rows(min_row=ws2.max_row - 4, max_row=ws2.max_row, values_only=False):
    for cell in row:
        if cell.column_letter in ("H", "I"):
            print(f"  Row {cell.row}, Col {cell.column_letter}: value={repr(cell.value)}, type={type(cell.value).__name__}")

wb.close()
wb2.close()
