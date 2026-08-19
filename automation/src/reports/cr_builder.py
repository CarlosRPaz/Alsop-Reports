"""
cr_builder.py — Close Rate Report builder.

Generates a Close Rate report with two sheets:
  1. "Auto YTD" — Year-to-date close rate per agent
  2. "Auto Monthly" — Current month close rate per agent

Columns (matching original CR sheet structure):
  A: Sub Producer (agent name)
  B: Issued Pol. Cnt (NB count)
  C: Quote Cnt
  D: Close Rate (%)
  E: Auto Items
  F: Monthly quotes for 30 autos (= (30/1.25)/CloseRate)
  G: Daily Quote Goal (= MonthlyQuotes/21)
  H: Daily Quote-Actual (= QuoteCnt / BusinessDaysPassed)
  I: Team (hidden)
  J: Office (hidden)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import date, timedelta
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from src.spine import Spine


# ── Style constants (matching original CR sheet — Aptos Narrow) ─────────────
_FONT_TITLE = Font(name="Aptos Narrow", size=16, bold=True)
_FONT_HEADER = Font(name="Aptos Narrow", size=11, bold=True)
_FONT_DATA = Font(name="Aptos Narrow", size=11)
_FONT_TOTAL = Font(name="Aptos Narrow", size=11, bold=True)
_FONT_NOTE_RED = Font(name="Aptos Narrow", size=11, color="9C0006")
_FONT_NOTE_GREEN = Font(name="Aptos Narrow", size=11, color="006100")
_FONT_LABEL = Font(name="Aptos Narrow", size=11, bold=True)
_FONT_DATE = Font(name="Aptos Narrow", size=11, bold=True, color="C00000")
_FONT_ACTUAL_HDR = Font(name="Aptos Narrow", size=11, color="9C0006")

_FILL_PINK = PatternFill("solid", fgColor="FFC7CE")     # Warning/note
_FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")    # On pace indicator
_FILL_GOLD = PatternFill("solid", fgColor="FFC000")     # Production header

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

_MED_SIDE = Side(style="medium")

# Column definitions for CR table
CR_COLUMNS = [
    ("Sub Producer",                 29,   None),
    ("Issued Pol. Cnt",              13,   "#,##0"),
    ("Quote Cnt",                    13,   "#,##0"),
    ("Close Rate",                   13,   "0.00%"),
    ("Auto Items",                    9,   "#,##0"),
    ("Monthly quotes for 30 autos", 10,   "0"),
    ("Daily Quote Goal",              9,   "0.0"),
    ("Daily Quote-Actual",            9,   "0.0"),
    ("Team",                          8,   None),
    ("Office",                       18,   None),
]


def _business_days_in_range(start: date, end: date) -> int:
    """Count business days (Mon-Fri) between start and end, inclusive."""
    count = 0
    current = start
    while current <= end:
        if current.weekday() < 5:  # Mon=0, Fri=4
            count += 1
        current += timedelta(days=1)
    return count


def _business_days_ytd(target_date: date) -> int:
    """Business days from Jan 1 through target_date."""
    year_start = date(target_date.year, 1, 1)
    return _business_days_in_range(year_start, target_date)


def _business_days_mtd(target_date: date) -> int:
    """Business days from 1st of month through target_date."""
    month_start = target_date.replace(day=1)
    return _business_days_in_range(month_start, target_date)


def build_cr_from_dsr(
    dsr_path: str,
    spine: Spine,
    target_date: date,
    output_folder: str = "reports",
) -> Path:
    """
    Build a Close Rate report with YTD and Monthly sheets.

    Reads Quotes and NB data from the DSR master workbook.
    """
    # Read DSR data
    df = pd.read_excel(dsr_path, engine="openpyxl", sheet_name="DSR", header=2)
    df.columns = [str(c).strip() for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])

    # Fill missing columns
    for col in ["Quotes", "NB", "Items", "Total Premium", "Team", "Office", "Agent"]:
        if col not in df.columns:
            df[col] = 0
        else:
            df[col] = df[col].fillna(0) if col not in ("Team", "Office", "Agent") else df[col].fillna("")

    wb = Workbook()

    # Sheet 1: Auto YTD
    year_start = date(target_date.year, 1, 1)
    ytd_mask = (df["Date"].dt.date >= year_start) & (df["Date"].dt.date <= target_date)
    ytd_df = df[ytd_mask]
    biz_days_ytd = _business_days_ytd(target_date)

    ytd_agg = _aggregate_cr(ytd_df)
    _write_cr_sheet(wb, wb.active, "Auto YTD", ytd_agg, biz_days_ytd,
                    target_date, "YTD", year_start, target_date)

    # Sheet 2: Auto Monthly
    month_start = target_date.replace(day=1)
    mtd_mask = (df["Date"].dt.date >= month_start) & (df["Date"].dt.date <= target_date)
    mtd_df = df[mtd_mask]
    biz_days_mtd = _business_days_mtd(target_date)

    ws_monthly = wb.create_sheet("Auto Monthly")
    monthly_agg = _aggregate_cr(mtd_df)
    _write_cr_sheet(wb, ws_monthly, "Auto Monthly", monthly_agg, biz_days_mtd,
                    target_date, "Monthly", month_start, target_date)

    # Save
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    filename = output_path / f"CR_{target_date.isoformat()}.xlsx"
    wb.save(str(filename))
    print(f"[cr_builder] Report saved: {filename}")
    return filename


def _aggregate_cr(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate quotes and NB per agent for close rate calculation."""
    if len(df) == 0:
        return pd.DataFrame(columns=["Agent", "Team", "Office", "NB", "Quotes", "Items"])

    agg = df.groupby("Agent").agg(
        Team=("Team", "first"),
        Office=("Office", "first"),
        NB=("NB", "sum"),
        Quotes=("Quotes", "sum"),
        Items=("Items", "sum"),
    ).reset_index()

    # Filter out agents with no activity
    agg = agg[(agg["NB"] > 0) | (agg["Quotes"] > 0)]
    agg = agg.sort_values("Agent")
    return agg


def _write_cr_sheet(wb, ws, title: str, agg: pd.DataFrame, biz_days: int,
                     target_date: date, period: str,
                     period_start: date, period_end: date):
    """Write a Close Rate sheet (YTD or Monthly)."""
    ws.title = title
    ws.sheet_view.zoomScale = 100

    # ── Row 1: Info header ──
    ws["A1"] = f"Most Recent Data Through"
    ws["A1"].font = _FONT_LABEL
    ws["B1"] = period_end
    ws["B1"].font = _FONT_DATE
    ws["B1"].number_format = "mm/dd/yy"

    # ── Row 3: Note about duplicate quotes ──
    ws.merge_cells("C3:E4")
    ws["C3"] = "Close rate based on raw quote count from DSR data."
    ws["C3"].font = _FONT_NOTE_RED
    ws["C3"].fill = _FILL_PINK
    ws["C3"].alignment = Alignment(wrap_text=True, vertical="center")

    # ── Row 5: Business days ──
    ws["A5"] = f"Business Days Passed ({period})"
    ws["A5"].font = _FONT_LABEL
    ws["A6"] = biz_days
    ws["A6"].font = _FONT_DATA

    # ── Row 7: On-pace indicator ──
    ws["C7"] = "On pace"
    ws["C7"].font = _FONT_NOTE_GREEN
    ws["C7"].fill = _FILL_GREEN
    ws["D7"] = ">=15%"
    ws["D7"].font = _FONT_NOTE_GREEN
    ws["D7"].fill = _FILL_GREEN

    # ── Row 8: Section title ──
    ws.merge_cells("A8:H8")
    ws["A8"] = title
    ws["A8"].font = _FONT_TITLE

    # ── Row 9: Totals ──
    totals_row = 9
    ws.cell(row=totals_row, column=1, value="Total").font = _FONT_TOTAL

    total_nb = int(agg["NB"].sum()) if len(agg) > 0 else 0
    total_quotes = int(agg["Quotes"].sum()) if len(agg) > 0 else 0
    total_cr = total_nb / total_quotes if total_quotes > 0 else 0
    total_items = int(agg["Items"].sum()) if len(agg) > 0 else 0

    ws.cell(row=totals_row, column=2, value=total_nb).font = _FONT_TOTAL
    ws.cell(row=totals_row, column=3, value=total_quotes).font = _FONT_TOTAL
    ws.cell(row=totals_row, column=4, value=total_cr).font = _FONT_TOTAL
    ws.cell(row=totals_row, column=4).number_format = "0.00%"

    if biz_days > 0:
        total_actual = total_quotes / biz_days
        ws.cell(row=totals_row, column=8, value=round(total_actual, 1)).font = _FONT_TOTAL
        ws.cell(row=totals_row, column=8).number_format = "0.0"

    # ── Row 10: Column headers ──
    header_row = 10
    for col_idx, (header, width, fmt) in enumerate(CR_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Daily Quote-Actual header in red
    ws.cell(row=header_row, column=8).font = _FONT_ACTUAL_HDR
    ws.cell(row=header_row, column=8).fill = _FILL_PINK

    # Hide Team and Office columns
    ws.column_dimensions["I"].hidden = True
    ws.column_dimensions["J"].hidden = True

    # ── Data rows ──
    data_start = header_row + 1
    for row_idx, (_, row) in enumerate(agg.iterrows()):
        r = data_start + row_idx
        agent = row["Agent"]
        nb = int(row["NB"])
        quotes = int(row["Quotes"])
        close_rate = nb / quotes if quotes > 0 else 0
        items = int(row["Items"])

        # Monthly quotes for 30 autos: (30/1.25) / close_rate
        monthly_30 = int(round((30 / 1.25) / close_rate)) if close_rate > 0 else 0
        # Daily quote goal = monthly_30 / 21
        daily_goal = round(monthly_30 / 21, 1) if monthly_30 > 0 else 0
        # Daily actual = quotes / business days
        daily_actual = round(quotes / biz_days, 1) if biz_days > 0 else 0

        values = [
            agent,
            nb,
            quotes,
            close_rate,
            items,
            monthly_30 if monthly_30 > 0 else "-",
            daily_goal if daily_goal > 0 else "-",
            daily_actual,
            row.get("Team", ""),
            row.get("Office", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = _FONT_DATA

            _, _, fmt = CR_COLUMNS[col_idx - 1]
            if fmt and not isinstance(val, str):
                cell.number_format = fmt

    # ── Conditional formatting ──
    last_row = data_start + len(agg) - 1
    if last_row >= data_start:
        # Close rate >= 15% = green fill
        ws.conditional_formatting.add(
            f"D{data_start}:D{last_row}",
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["0.15"],
                font=Font(name="Aptos Narrow", size=11, color="006100"),
                fill=_FILL_GREEN,
            ),
        )

        # Close rate < 15% and > 0 = pink
        ws.conditional_formatting.add(
            f"D{data_start}:D{last_row}",
            CellIsRule(
                operator="lessThan",
                formula=["0.15"],
                font=Font(name="Aptos Narrow", size=11, color="9C0006"),
                fill=_FILL_PINK,
            ),
        )

        # Zeros = gray
        ws.conditional_formatting.add(
            f"B{data_start}:H{last_row}",
            CellIsRule(
                operator="equal",
                formula=["0"],
                font=Font(color="A6A6A6"),
            ),
        )

    # Freeze panes
    ws.freeze_panes = f"A{data_start}"
