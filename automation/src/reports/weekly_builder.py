"""
weekly_builder.py — Weekly Production Report builder.

Aggregates DSR daily data for a Mon-Sun week into a styled Excel report.
Matches the format of the existing Weekly sheet in the DSR workbook.

Columns (matching original Weekly sheet):
  A: (hidden/unused)
  B: (date ref, hidden)
  C: Office
  D: Name (Agent)
  E: In Calls
  F: Out Calls
  G: Total Calls
  H: Talk Time
  I: Texts
  J: Unique Leads       (from Rico screenshots — manual/OCR)
  K: Rico Hot Pipeline   (from Rico screenshots — manual/OCR)
  L: #PIVOT              (from Rico screenshots — manual/OCR)
  M: #SAVED              (from Rico screenshots — manual/OCR)
  N: eAgent Dismissed To-do's
  O: eAgent Past Due To-Do's
  P: Rico Past Due Tasks (manual)
  Q: Life Leads          (manual)
  R: Bundle MTD          (manual)
  S: Bundle YTD          (manual)
  T: Auto Quotes
  U: Total Written Premium Wk
  V: MTD Total Premium
  W: Auto Pts Wk         (from Premium/AgencyZoom)
  X: Prev Mo Auto Pts    (manual)
  Y: MTD Auto Items
"""

import pandas as pd
from pathlib import Path
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.spine import Spine


# ── Style constants ──────────────────────────────────────────────────────────
_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_TITLE = Font(name="Calibri", size=14, bold=True)
_FONT_DATA = Font(name="Calibri", size=10)
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

_FILL_GREEN = PatternFill("solid", fgColor="4472C4")    # Blue header for calls
_FILL_PURPLE = PatternFill("solid", fgColor="7030A0")   # Purple for Rico/eAgent
_FILL_GOLD = PatternFill("solid", fgColor="BF8F00")     # Gold for production
_FILL_TOTALS = PatternFill("solid", fgColor="4EA72E")   # Green totals row
_FILL_LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")  # Alternating rows

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(bottom=_THIN_SIDE)

# Column definitions: (header, width, number_format, fill_group)
WEEKLY_COLUMNS = [
    ("Office",                    8,  None,           "calls"),
    ("Name",                     14,  None,           "calls"),
    ("In Calls",                  9,  "#,##0",        "calls"),
    ("Out Calls",                10,  "#,##0",        "calls"),
    ("Total Calls",              11,  "#,##0",        "calls"),
    ("Talk Time",                11,  "[h]:mm:ss",    "calls"),
    ("Texts",                     8,  "#,##0",        "calls"),
    ("Unique Leads",             12,  "#,##0",        "rico"),
    ("Rico Hot Pipeline",        15,  "#,##0",        "rico"),
    ("#PIVOT",                    8,  "#,##0",        "rico"),
    ("#SAVED",                    8,  "#,##0",        "rico"),
    ("Dismissed To-Do's",        16,  "#,##0",        "eagent"),
    ("Past Due To-Do's",         15,  "#,##0",        "eagent"),
    ("Rico Past Due Tasks",      16,  "#,##0",        "rico"),
    ("Life Leads",               10,  "#,##0",        "rico"),
    ("Bundle MTD",               10,  "#,##0",        "production"),
    ("Bundle YTD",               10,  "#,##0",        "production"),
    ("Auto Quotes",              11,  "#,##0",        "production"),
    ("Total Written Premium Wk", 20,  '"$"#,##0',     "production"),
    ("MTD Total Premium",        16,  '"$"#,##0',     "production"),
    ("Auto Pts Wk",              11,  "#,##0",        "production"),
    ("Prev Mo Auto Pts",         14,  "#,##0",        "production"),
    ("MTD Auto Items",           13,  "#,##0",        "production"),
]

_FILL_MAP = {
    "calls": _FILL_GREEN,
    "rico": _FILL_PURPLE,
    "eagent": _FILL_PURPLE,
    "production": _FILL_GOLD,
}


def get_week_range(target_date: date) -> tuple[date, date]:
    """Return (Monday, Sunday) for the week containing target_date."""
    monday = target_date - timedelta(days=target_date.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def get_week_label(monday: date, sunday: date) -> str:
    """Format week label like '03.30-04.05'."""
    return f"{monday.strftime('%m.%d')}-{sunday.strftime('%m.%d')}"


def aggregate_weekly(dsr_data: pd.DataFrame, week_start: date, week_end: date,
                     spine: Spine) -> pd.DataFrame:
    """
    Aggregate DSR daily rows into weekly totals per agent.

    Parameters
    ----------
    dsr_data : DataFrame
        Raw DSR data with columns: Date, Agent, Team, Office, Calls, Inbound, Outbound,
        TalkTimeSeconds, Texts, OutTexts, Quotes, NB, Items, Total Premium, etc.
    week_start : date
        Monday of the target week.
    week_end : date
        Sunday of the target week.
    spine : Spine
        Agent name resolver (for metadata).

    Returns
    -------
    DataFrame with one row per agent, weekly totals.
    """
    df = dsr_data.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    mask = (df["Date"].dt.date >= week_start) & (df["Date"].dt.date <= week_end)
    df = df[mask]

    if len(df) == 0:
        return pd.DataFrame()

    # Aggregate per agent
    agg = df.groupby("Agent").agg(
        Office=("Office", "first"),
        Team=("Team", "first"),
        InCalls=("Inbound", "sum"),
        OutCalls=("Outbound", "sum"),
        TotalCalls=("Calls", "sum"),
        TalkTimeSeconds=("TalkTimeSeconds", lambda x: x.sum() if "TalkTimeSeconds" in df.columns else 0),
        Texts=("Texts", "sum"),
        Quotes=("Quotes", "sum"),
        NB=("NB", "sum"),
        Items=("Items", "sum"),
        TotalPremium=("Total Premium", "sum"),
        DismissedTodos=("Dismissed To-Do's", "last"),  # Last day's value (cumulative)
        PastDueTodos=("Past Due To-Do's", "last"),
        PivotCount=("pivot_count", "sum"),
        Contact=("Contact", "sum"),
        Quoted_pipeline=("Quoted", "sum"),
        Hot=("Hot", "sum"),
        Xsale=("XDate", "sum"),
    ).reset_index()

    return agg.sort_values(["Team", "Agent"])


def build_weekly_from_dsr(
    dsr_path: str,
    spine: Spine,
    target_date: date,
    output_folder: str = "reports",
) -> Path:
    """
    Build a weekly report by reading existing DSR data from the master workbook.

    Parameters
    ----------
    dsr_path : str
        Path to the DSR master workbook (or original).
    spine : Spine
        Agent name resolver.
    target_date : date
        Any date within the target week (will find Mon-Sun).
    output_folder : str
        Where to save the output.

    Returns
    -------
    Path to the generated weekly report.
    """
    week_start, week_end = get_week_range(target_date)
    week_label = get_week_label(week_start, week_end)

    # Read DSR data
    df = pd.read_excel(dsr_path, engine="openpyxl", sheet_name="DSR", header=2)
    df.columns = [str(c).strip() for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])

    # Convert Talk Time to seconds for aggregation
    if "Talk Time" in df.columns:
        def _to_seconds(val):
            if pd.isna(val):
                return 0
            if isinstance(val, timedelta):
                return int(val.total_seconds())
            s = str(val)
            parts = s.split(":")
            try:
                if len(parts) == 3:
                    return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                elif len(parts) == 2:
                    return int(parts[0]) * 60 + int(parts[1])
            except ValueError:
                return 0
            return 0
        df["TalkTimeSeconds"] = df["Talk Time"].apply(_to_seconds)
    else:
        df["TalkTimeSeconds"] = 0

    # Fill missing columns
    for col in ["Calls", "Inbound", "Outbound", "Texts", "Quotes", "NB", "Items",
                "Total Premium", "Contact", "Quoted", "Hot", "XDate",
                "Dismissed To-Do's", "Past Due To-Do's", "pivot_count"]:
        if col not in df.columns:
            df[col] = 0
        else:
            df[col] = df[col].fillna(0)

    # Aggregate
    agg = aggregate_weekly(df, week_start, week_end, spine)

    if len(agg) == 0:
        print(f"[weekly_builder] No data for week {week_label}")
        return Path(output_folder) / f"Weekly_{week_label}.xlsx"

    # Calculate MTD premium (from month start through week end)
    month_start = week_end.replace(day=1)
    mtd_mask = (df["Date"].dt.date >= month_start) & (df["Date"].dt.date <= week_end)
    mtd_df = df[mtd_mask]
    mtd_premium = mtd_df.groupby("Agent")["Total Premium"].sum().to_dict()
    mtd_items = mtd_df.groupby("Agent")["Items"].sum().to_dict()

    # Build the Excel report
    return _write_weekly_excel(agg, week_label, week_start, week_end,
                               mtd_premium, mtd_items, output_folder)


def _write_weekly_excel(
    agg: pd.DataFrame,
    week_label: str,
    week_start: date,
    week_end: date,
    mtd_premium: dict,
    mtd_items: dict,
    output_folder: str,
) -> Path:
    """Write the styled weekly Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly"
    ws.sheet_view.zoomScale = 90

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Week: {week_label}"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("H1:N1")
    ws["H1"] = "Weekly Production"
    ws["H1"].font = Font(name="Calibri", size=14, bold=True, color="7030A0")

    # Row 2: Subtitle
    ws["A2"] = "Mon - Sun"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    # Row 3: Headers
    header_row = 3
    for col_idx, (header, width, fmt, group) in enumerate(WEEKLY_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_MAP.get(group, _FILL_GREEN)
        cell.alignment = _ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    data_start = header_row + 1
    for row_idx, (_, row) in enumerate(agg.iterrows()):
        r = data_start + row_idx
        agent = row["Agent"]
        talk_seconds = int(row.get("TalkTimeSeconds", 0))
        talk_fraction = talk_seconds / 86400.0 if talk_seconds > 0 else 0

        values = [
            row.get("Office", ""),
            agent,
            int(row.get("InCalls", 0)),
            int(row.get("OutCalls", 0)),
            int(row.get("TotalCalls", 0)),
            talk_fraction,  # Excel time fraction
            int(row.get("Texts", 0)),
            0,  # Unique Leads (manual/screenshot)
            int(row.get("Hot", 0)),  # Rico Hot Pipeline
            int(row.get("PivotCount", 0)),  # #PIVOT (eAgent/screenshot)
            0,  # #SAVED (manual/screenshot)
            int(row.get("DismissedTodos", 0)),
            int(row.get("PastDueTodos", 0)),
            0,  # Rico Past Due Tasks (manual)
            0,  # Life Leads (manual)
            0,  # Bundle MTD (manual)
            0,  # Bundle YTD (manual)
            int(row.get("Quotes", 0)),
            float(row.get("TotalPremium", 0)),
            float(mtd_premium.get(agent, 0)),
            0,  # Auto Pts Wk (from Premium/AgencyZoom)
            0,  # Prev Mo Auto Pts (manual)
            int(mtd_items.get(agent, 0)),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = _FONT_DATA

            # Apply number format
            _, _, fmt, _ = WEEKLY_COLUMNS[col_idx - 1]
            if fmt:
                cell.number_format = fmt

            # Alternating row fill
            if row_idx % 2 == 0:
                cell.border = _BORDER

    # Totals row
    last_data_row = data_start + len(agg) - 1
    totals_row = last_data_row + 1
    ws.cell(row=totals_row, column=1, value="TOTALS").font = _FONT_TOTAL
    ws.cell(row=totals_row, column=1).fill = _FILL_TOTALS

    for col_idx in range(1, len(WEEKLY_COLUMNS) + 1):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = _FILL_TOTALS
        cell.font = _FONT_TOTAL
        _, _, fmt, _ = WEEKLY_COLUMNS[col_idx - 1]
        if fmt and col_idx >= 3:  # Skip Office/Name columns
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{last_data_row})"
            if fmt:
                cell.number_format = fmt

    # Freeze panes
    ws.freeze_panes = "C4"

    # Save
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    filename = output_path / f"Weekly_{week_label.replace('.', '').replace('-', '_')}.xlsx"
    wb.save(str(filename))
    print(f"[weekly_builder] Report saved: {filename}")
    return filename
