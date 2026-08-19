"""
monthly_builder.py — Monthly Performance Report builder.

Aggregates DSR daily data for a full calendar month into a styled Excel report.
Since the original Monthly sheet was "UNDER CONSTRUCTION", this creates a
comprehensive monthly view with:
  - Agent-level monthly totals for all metrics
  - Daily averages
  - Close rate (NB / Quotes)
  - Team subtotals

Two sheets:
  1. "Monthly Summary" — one row per agent, full month totals + averages
  2. "Daily Breakdown" — agent x date matrix showing daily production
"""

import pandas as pd
from pathlib import Path
from datetime import date, timedelta
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.spine import Spine


# ── Style constants ──────────────────────────────────────────────────────────
_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_TITLE = Font(name="Calibri", size=14, bold=True)
_FONT_DATA = Font(name="Calibri", size=10)
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_SUBTOTAL = Font(name="Calibri", size=10, bold=True)

_FILL_BLUE = PatternFill("solid", fgColor="4472C4")
_FILL_PURPLE = PatternFill("solid", fgColor="7030A0")
_FILL_GOLD = PatternFill("solid", fgColor="BF8F00")
_FILL_GREEN = PatternFill("solid", fgColor="4EA72E")
_FILL_DARK_GREEN = PatternFill("solid", fgColor="375623")
_FILL_LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
_FILL_CSR = PatternFill("solid", fgColor="E2EFDA")
_FILL_SALES = PatternFill("solid", fgColor="FCE4D6")
_FILL_EA = PatternFill("solid", fgColor="D6DCE4")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BORDER = Border(bottom=_THIN_SIDE)

# Monthly summary columns
MONTHLY_COLUMNS = [
    # (header, width, format, header_fill)
    ("Team",                  7,   None,           _FILL_BLUE),
    ("Office",                8,   None,           _FILL_BLUE),
    ("Agent",                14,   None,           _FILL_BLUE),
    ("Days Worked",           10,  "#,##0",        _FILL_BLUE),
    ("Total Calls",           11,  "#,##0",        _FILL_BLUE),
    ("Avg Calls/Day",         12,  "#,##0.0",      _FILL_BLUE),
    ("Inbound",               9,   "#,##0",        _FILL_BLUE),
    ("Outbound",              10,  "#,##0",        _FILL_BLUE),
    ("Talk Time",             11,  "[h]:mm:ss",    _FILL_BLUE),
    ("Avg Talk/Day",          11,  "[h]:mm:ss",    _FILL_BLUE),
    ("Texts",                 8,   "#,##0",        _FILL_PURPLE),
    ("Avg Texts/Day",         12,  "#,##0.0",      _FILL_PURPLE),
    ("Out Texts",             10,  "#,##0",        _FILL_PURPLE),
    ("Opt-Ins",               8,   "#,##0",        _FILL_PURPLE),
    ("Opt-Outs",              9,   "#,##0",        _FILL_PURPLE),
    ("Quotes",                8,   "#,##0",        _FILL_GOLD),
    ("NB",                    5,   "#,##0",        _FILL_GOLD),
    ("Close Rate",            10,  "0.0%",         _FILL_GOLD),
    ("Items",                 7,   "#,##0",        _FILL_GOLD),
    ("Total Premium",         14,  '"$"#,##0',     _FILL_GOLD),
    ("Avg Premium/Day",       14,  '"$"#,##0',     _FILL_GOLD),
    ("Contact",               8,   "#,##0",        _FILL_PURPLE),
    ("Quoted",                8,   "#,##0",        _FILL_PURPLE),
    ("Hot",                   5,   "#,##0",        _FILL_PURPLE),
    ("XDate",                7,   "#,##0",        _FILL_PURPLE),
]

_TEAM_FILLS = {"CSR": _FILL_CSR, "Sales": _FILL_SALES, "EA": _FILL_EA}


def build_monthly_from_dsr(
    dsr_path: str,
    spine: Spine,
    target_month: date,
    output_folder: str = "reports",
) -> Path:
    """
    Build a monthly report from DSR data.

    Parameters
    ----------
    dsr_path : str
        Path to the DSR master workbook.
    spine : Spine
        Agent name resolver.
    target_month : date
        Any date within the target month.
    output_folder : str
        Where to save the output.
    """
    month_start = target_month.replace(day=1)
    _, last_day = monthrange(target_month.year, target_month.month)
    month_end = date(target_month.year, target_month.month, last_day)
    month_label = month_start.strftime("%B %Y")  # e.g. "March 2026"

    # Read DSR data
    df = pd.read_excel(dsr_path, engine="openpyxl", sheet_name="DSR", header=2)
    df.columns = [str(c).strip() for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])

    # Filter to target month
    mask = (df["Date"].dt.date >= month_start) & (df["Date"].dt.date <= month_end)
    df = df[mask]

    if len(df) == 0:
        print(f"[monthly_builder] No data for {month_label}")
        return Path(output_folder) / f"Monthly_{month_start.strftime('%Y-%m')}.xlsx"

    # Convert Talk Time to seconds
    if "Talk Time" in df.columns:
        def _to_seconds(val):
            if pd.isna(val):
                return 0
            if isinstance(val, timedelta):
                return int(val.total_seconds())
            s = str(val)
            parts = s.split(":")
            try:
                if len(parts) == 3:
                    return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                elif len(parts) == 2:
                    return int(parts[0]) * 60 + int(parts[1])
            except ValueError:
                return 0
            return 0
        df["TalkTimeSeconds"] = df["Talk Time"].apply(_to_seconds)
    else:
        df["TalkTimeSeconds"] = 0

    # Fill missing columns
    for col in ["Calls", "Inbound", "Outbound", "Texts", "OutTexts", "Opt-Ins", "Opt-Outs",
                "Quotes", "NB", "Items", "Total Premium", "Contact", "Quoted", "Hot", "XDate",
                "Dismissed To-Do's", "Past Due To-Do's"]:
        if col not in df.columns:
            df[col] = 0
        else:
            df[col] = df[col].fillna(0)

    # Aggregate per agent
    agg = df.groupby("Agent").agg(
        Team=("Team", "first"),
        Office=("Office", "first"),
        DaysWorked=("Date", "nunique"),
        TotalCalls=("Calls", "sum"),
        Inbound=("Inbound", "sum"),
        Outbound=("Outbound", "sum"),
        TalkTimeSeconds=("TalkTimeSeconds", "sum"),
        Texts=("Texts", "sum"),
        OutTexts=("OutTexts", "sum"),
        OptIns=("Opt-Ins", "sum"),
        OptOuts=("Opt-Outs", "sum"),
        Quotes=("Quotes", "sum"),
        NB=("NB", "sum"),
        Items=("Items", "sum"),
        TotalPremium=("Total Premium", "sum"),
        Contact=("Contact", "sum"),
        Quoted_pipeline=("Quoted", "sum"),
        Hot=("Hot", "sum"),
        Xsale=("XDate", "sum"),
    ).reset_index()

    # Derived columns
    agg["AvgCalls"] = (agg["TotalCalls"] / agg["DaysWorked"]).round(1)
    agg["AvgTalkSeconds"] = (agg["TalkTimeSeconds"] / agg["DaysWorked"]).astype(int)
    agg["AvgTexts"] = (agg["Texts"] / agg["DaysWorked"]).round(1)
    agg["CloseRate"] = (agg["NB"] / agg["Quotes"].replace(0, float("nan"))).fillna(0)
    agg["AvgPremium"] = (agg["TotalPremium"] / agg["DaysWorked"]).round(0)

    agg = agg.sort_values(["Team", "Agent"])

    # Build workbook
    wb = Workbook()
    _write_summary_sheet(wb, agg, month_label, month_start, month_end)
    _write_daily_breakdown(wb, df, month_label, month_start, month_end)

    # Save
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    filename = output_path / f"Monthly_{month_start.strftime('%Y-%m')}.xlsx"
    wb.save(str(filename))
    print(f"[monthly_builder] Report saved: {filename}")
    return filename


def _write_summary_sheet(wb: Workbook, agg: pd.DataFrame, month_label: str,
                          month_start: date, month_end: date):
    """Write the Monthly Summary sheet."""
    ws = wb.active
    ws.title = "Monthly Summary"
    ws.sheet_view.zoomScale = 90

    # Row 1: Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Monthly Performance Report - {month_label}"
    ws["A1"].font = _FONT_TITLE

    # Row 2: Date range
    ws["A2"] = f"{month_start.strftime('%m/%d/%Y')} - {month_end.strftime('%m/%d/%Y')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    # Row 3: Headers
    header_row = 3
    for col_idx, (header, width, fmt, fill) in enumerate(MONTHLY_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = fill
        cell.alignment = _ALIGN_WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows, grouped by team
    data_start = header_row + 1
    r = data_start
    current_team = None

    for _, row in agg.iterrows():
        team = row["Team"]

        # Team subtotal separator
        if team != current_team and current_team is not None:
            # Insert team subtotal for previous team
            r = _write_team_subtotal(ws, r, data_start, current_team, agg, MONTHLY_COLUMNS)
            r += 1  # blank row
            data_start = r  # reset for next team

        current_team = team
        talk_fraction = row["TalkTimeSeconds"] / 86400.0
        avg_talk_fraction = row["AvgTalkSeconds"] / 86400.0

        values = [
            team,
            row.get("Office", ""),
            row["Agent"],
            int(row["DaysWorked"]),
            int(row["TotalCalls"]),
            row["AvgCalls"],
            int(row["Inbound"]),
            int(row["Outbound"]),
            talk_fraction,
            avg_talk_fraction,
            int(row["Texts"]),
            row["AvgTexts"],
            int(row["OutTexts"]),
            int(row["OptIns"]),
            int(row["OptOuts"]),
            int(row["Quotes"]),
            int(row["NB"]),
            row["CloseRate"],
            int(row["Items"]),
            float(row["TotalPremium"]),
            float(row["AvgPremium"]),
            int(row.get("Contact", 0)),
            int(row.get("Quoted_pipeline", 0)),
            int(row.get("Hot", 0)),
            int(row.get("Xsale", 0)),
        ]

        team_fill = _TEAM_FILLS.get(team)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = _FONT_DATA
            _, _, fmt, _ = MONTHLY_COLUMNS[col_idx - 1]
            if fmt:
                cell.number_format = fmt
            if team_fill and col_idx <= 3:  # Color team/office/agent cells
                cell.fill = team_fill

        r += 1

    # Final team subtotal
    if current_team is not None:
        r = _write_team_subtotal(ws, r, data_start, current_team, agg, MONTHLY_COLUMNS)

    # Grand totals
    r += 1
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = _FONT_TOTAL
    for col_idx in range(1, len(MONTHLY_COLUMNS) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = _FILL_DARK_GREEN
        cell.font = _FONT_TOTAL

    # Sum numeric columns for grand total
    _set_grand_totals(ws, r, agg, MONTHLY_COLUMNS)

    # Freeze
    ws.freeze_panes = "D4"


def _write_team_subtotal(ws, r, team_data_start, team, agg, columns) -> int:
    """Write a team subtotal row. Returns next row number."""
    ws.cell(row=r, column=1, value=f"{team} Total").font = _FONT_SUBTOTAL
    team_data = agg[agg["Team"] == team]

    _set_team_totals(ws, r, team_data, columns)

    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=r, column=col_idx).fill = _FILL_LIGHT_GRAY
        ws.cell(row=r, column=col_idx).font = _FONT_SUBTOTAL

    return r + 1


def _set_team_totals(ws, r, team_data, columns):
    """Set team subtotal values."""
    col_map = {
        3: ("DaysWorked", "sum"),   # Days Worked
        4: ("TotalCalls", "sum"),
        5: ("AvgCalls", "mean"),
        6: ("Inbound", "sum"),
        7: ("Outbound", "sum"),
        8: ("TalkTimeSeconds", "sum"),    # Talk Time (seconds -> fraction)
        9: ("AvgTalkSeconds", "mean"),    # Avg Talk (seconds -> fraction)
        10: ("Texts", "sum"),
        11: ("AvgTexts", "mean"),
        12: ("OutTexts", "sum"),
        13: ("OptIns", "sum"),
        14: ("OptOuts", "sum"),
        15: ("Quotes", "sum"),
        16: ("NB", "sum"),
        17: None,  # Close Rate — calculated
        18: ("Items", "sum"),
        19: ("TotalPremium", "sum"),
        20: ("AvgPremium", "mean"),
        21: ("Contact", "sum"),
        22: ("Quoted_pipeline", "sum"),
        23: ("Hot", "sum"),
        24: ("Xsale", "sum"),
    }

    for col_idx, mapping in col_map.items():
        if mapping is None:
            # Close rate
            total_quotes = team_data["Quotes"].sum()
            total_nb = team_data["NB"].sum()
            val = total_nb / total_quotes if total_quotes > 0 else 0
        else:
            col_name, agg_func = mapping
            if col_name in team_data.columns:
                val = getattr(team_data[col_name], agg_func)()
            else:
                val = 0

        # Convert seconds to Excel time fraction for time columns
        if col_idx in (8, 9):
            val = val / 86400.0

        cell = ws.cell(row=r, column=col_idx + 1, value=val)
        _, _, fmt, _ = columns[col_idx]
        if fmt:
            cell.number_format = fmt


def _set_grand_totals(ws, r, agg, columns):
    """Set grand total values."""
    total_quotes = agg["Quotes"].sum()
    total_nb = agg["NB"].sum()

    values = {
        3: int(agg["DaysWorked"].sum()),
        4: int(agg["TotalCalls"].sum()),
        5: agg["AvgCalls"].mean(),
        6: int(agg["Inbound"].sum()),
        7: int(agg["Outbound"].sum()),
        8: agg["TalkTimeSeconds"].sum() / 86400.0,
        9: agg["AvgTalkSeconds"].mean() / 86400.0,
        10: int(agg["Texts"].sum()),
        11: agg["AvgTexts"].mean(),
        12: int(agg["OutTexts"].sum()),
        13: int(agg["OptIns"].sum()),
        14: int(agg["OptOuts"].sum()),
        15: int(total_quotes),
        16: int(total_nb),
        17: total_nb / total_quotes if total_quotes > 0 else 0,
        18: int(agg["Items"].sum()),
        19: float(agg["TotalPremium"].sum()),
        20: agg["AvgPremium"].mean(),
        21: int(agg["Contact"].sum()),
        22: int(agg["Quoted_pipeline"].sum()),
        23: int(agg["Hot"].sum()),
        24: int(agg["Xsale"].sum()),
    }

    for col_idx, val in values.items():
        cell = ws.cell(row=r, column=col_idx + 1, value=val)
        _, _, fmt, _ = columns[col_idx]
        if fmt:
            cell.number_format = fmt


def _write_daily_breakdown(wb: Workbook, df: pd.DataFrame, month_label: str,
                            month_start: date, month_end: date):
    """Write the Daily Breakdown sheet — agent rows x date columns."""
    ws = wb.create_sheet("Daily Breakdown")
    ws.sheet_view.zoomScale = 85

    # Get all dates in the month that have data
    dates_in_month = sorted(df["Date"].dt.date.unique())

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Daily Breakdown - {month_label}"
    ws["A1"].font = _FONT_TITLE

    # Headers: Agent | Team | then one column per date
    header_row = 3
    ws.cell(row=header_row, column=1, value="Agent").font = _FONT_HEADER
    ws.cell(row=header_row, column=1).fill = _FILL_BLUE
    ws.column_dimensions["A"].width = 14

    ws.cell(row=header_row, column=2, value="Team").font = _FONT_HEADER
    ws.cell(row=header_row, column=2).fill = _FILL_BLUE
    ws.column_dimensions["B"].width = 7

    ws.cell(row=header_row, column=3, value="Total").font = _FONT_HEADER
    ws.cell(row=header_row, column=3).fill = _FILL_GREEN
    ws.column_dimensions["C"].width = 8

    # Metric selector — show Calls by default
    ws["A2"] = "Metric: Total Calls"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, bold=True)

    for d_idx, d in enumerate(dates_in_month):
        col = 4 + d_idx
        cell = ws.cell(row=header_row, column=col, value=d.strftime("%m/%d"))
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.fill = _FILL_BLUE
        cell.alignment = _ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col)].width = 7

        # Day of week in row 2
        dow_cell = ws.cell(row=2, column=col, value=d.strftime("%a"))
        dow_cell.font = Font(name="Calibri", size=8, italic=True)
        dow_cell.alignment = _ALIGN_CENTER

    # Pivot: agent x date for Calls
    pivot = df.pivot_table(index="Agent", columns=df["Date"].dt.date,
                           values="Calls", aggfunc="sum", fill_value=0)

    agents = sorted([a for a in df["Agent"].unique() if pd.notna(a)])
    agent_teams = df.groupby("Agent")["Team"].first().to_dict()

    r = header_row + 1
    for agent in sorted(agents, key=lambda a: (agent_teams.get(a, ""), a)):
        team = agent_teams.get(agent, "")
        ws.cell(row=r, column=1, value=agent).font = _FONT_DATA
        ws.cell(row=r, column=2, value=team).font = _FONT_DATA

        # Team color
        team_fill = _TEAM_FILLS.get(team)
        if team_fill:
            ws.cell(row=r, column=1).fill = team_fill
            ws.cell(row=r, column=2).fill = team_fill

        # Total for the month
        total = int(pivot.loc[agent].sum()) if agent in pivot.index else 0
        ws.cell(row=r, column=3, value=total).font = Font(name="Calibri", size=10, bold=True)

        for d_idx, d in enumerate(dates_in_month):
            col = 4 + d_idx
            val = int(pivot.at[agent, d]) if agent in pivot.index and d in pivot.columns else 0
            cell = ws.cell(row=r, column=col, value=val if val > 0 else "")
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = _ALIGN_CENTER
        r += 1

    ws.freeze_panes = "D4"
