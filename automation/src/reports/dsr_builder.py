"""
dsr_builder.py — Daily Standup Report builder.

Combines data from all parsers + screenshot OCR into the formatted DSR layout,
matching the existing workbook's styling exactly.

DSR columns (matching existing workbook, starting at col B):
  B: Date  C: Team  D: Agent  E: Office  F: Full Name
  G: Calls  H: Inbound  I: Outbound  J: Talk Time
  K: Texts  L: OutTexts  M: Opt-Ins  N: Opt-Outs
  O: Quotes  P: NB  Q: Total Premium  R: Items
  S: MonthStart  T: Items MTD
  U: Contact  V: Quoted  W: Hot  X: XDate (Ricochet "3.3 XDATE - Task Set"; historically called "x-sale")
  Y: Dismissed To-Do's  Z: Past Due To-Do's
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, time
from pathlib import Path

from src.spine import Spine


# ===========================================================================
#  Style constants — matched from the existing DSR workbook
# ===========================================================================

# Core colors (from xl/theme/theme1.xml)
_BLACK = "000000"
_WHITE = "FFFFFF"
_DK2 = "0E2841"

# Header section fills (row 3)
_FILL_HDR_GRAY = PatternFill("solid", fgColor="BFC0C2")      # B-D (Date/Team/Agent)
_FILL_HDR_PURPLE = PatternFill("solid", fgColor="781F6E")     # K-N (Texts) — accent5 tint -0.25
_FILL_HDR_GOLD = PatternFill("solid", fgColor="FFC000")       # O-R, T (Quotes/NB/Premium)
_FILL_HDR_RED = PatternFill("solid", fgColor="FF552D")        # U-X (Rico pipeline)
_FILL_HDR_ACCENT = PatternFill("solid", fgColor="C97BB8")     # Y-Z (To-Do's) — accent5 tint +0.40

# Totals row fills (row 2)
_FILL_TOT_GREEN = PatternFill("solid", fgColor="4EA72E")      # G-J — accent6
_FILL_TOT_PURPLE = PatternFill("solid", fgColor="781F6E")     # K-N — accent5 tint -0.25
_FILL_TOT_GOLD = PatternFill("solid", fgColor="FFC000")       # O-R, T
_FILL_TOT_YELLOW = PatternFill("solid", fgColor="FFFF00")     # T2 static
_FILL_T1_LABEL = PatternFill("solid", fgColor="FFFFCC")       # T1 label
_FILL_EAGENT = PatternFill("solid", fgColor="F8D4C0")         # Y2:Z2 "eAgent only"

# Fonts
_FNT_TITLE = Font(name="Aptos Narrow", size=20, bold=True, color=_BLACK)
_FNT_A1 = Font(name="Aptos Narrow", size=11, bold=True, color="FF0000")
_FNT_T1 = Font(name="Aptos Narrow", size=12, bold=True, color=_BLACK)
_FNT_HDR_DEFAULT = Font(name="Aptos Narrow", size=11, color=_BLACK)
_FNT_HDR_PURPLE = Font(name="Aptos Narrow", size=12, bold=True, color=_WHITE)
_FNT_HDR_SMALL = Font(name="Aptos Narrow", size=10, color=_BLACK)
_FNT_TOT_WHITE = Font(name="Aptos Narrow", size=12, bold=True, color=_WHITE)
_FNT_TOT_DARK = Font(name="Aptos Narrow", size=12, bold=True, color=_BLACK)
_FNT_DATA = Font(name="Aptos Narrow", size=12, bold=True, color=_BLACK)
_FNT_DATA_SM = Font(name="Aptos Narrow", size=11, bold=True, color=_BLACK)
_FNT_EAGENT = Font(name="Aptos Narrow", size=10, bold=True, color=_BLACK)

# Border sides
_MED = Side(style="medium")
_THIN = Side(style="thin")
_NONE = Side()

# Column definitions
DSR_COLUMNS = [
    "Date", "Team", "Agent", "Office", "Full Name",
    "Calls", "Inbound", "Outbound", "Talk Time",
    "Texts", "OutTexts", "Opt-Ins", "Opt-Outs",
    "Quotes", "NB", "Total Premium", "Items",
    "MonthStart", "Items MTD",
    "Contact", "Quoted", "Hot", "XDate",
    "Dismissed To-Do's", "Past Due To-Do's",
]

# Column widths (exact from original)
COL_WIDTHS = {
    "A": 41.33,
    "B": 10.89, "C": 8.33, "D": 11.33, "E": 8.44, "F": 23.66,
    "G": 8.11, "H": 10.89, "I": 12.22, "J": 11.89,
    "K": 8.33, "L": 11.66, "M": 10.44, "N": 12.11,
    "O": 9.89, "P": 6.0, "Q": 16.33, "R": 8.66,
    "S": 13.0, "T": 12.78,
    "U": 10.33, "V": 10.0, "W": 6.66, "X": 8.89,
    "Y": 18.0, "Z": 16.56,
}

# Hidden columns
HIDDEN_COLS = ["E", "F", "S"]

# Header fill map (col letter -> fill)
_HDR_FILL = {
    "B": _FILL_HDR_GRAY, "C": _FILL_HDR_GRAY, "D": _FILL_HDR_GRAY,
    "E": _FILL_HDR_GRAY, "F": _FILL_HDR_GRAY,
    # G-J: no fill (inherits table style)
    "K": _FILL_HDR_PURPLE, "L": _FILL_HDR_PURPLE,
    "M": _FILL_HDR_PURPLE, "N": _FILL_HDR_PURPLE,
    "O": _FILL_HDR_GOLD, "P": _FILL_HDR_GOLD,
    "Q": _FILL_HDR_GOLD, "R": _FILL_HDR_GOLD,
    "T": _FILL_HDR_GOLD,
    "U": _FILL_HDR_RED, "V": _FILL_HDR_RED,
    "W": _FILL_HDR_RED, "X": _FILL_HDR_RED,
    "Y": _FILL_HDR_ACCENT, "Z": _FILL_HDR_ACCENT,
}

# Header font map (col letter -> font override)
_HDR_FONT = {
    "K": _FNT_HDR_PURPLE, "L": _FNT_HDR_PURPLE,
    "M": _FNT_HDR_PURPLE, "N": _FNT_HDR_PURPLE,
    "Y": _FNT_HDR_SMALL, "Z": _FNT_HDR_SMALL,
}

# Totals fill map
_TOT_FILL = {
    "G": _FILL_TOT_GREEN, "H": _FILL_TOT_GREEN,
    "I": _FILL_TOT_GREEN, "J": _FILL_TOT_GREEN,
    "K": _FILL_TOT_PURPLE, "L": _FILL_TOT_PURPLE,
    "M": _FILL_TOT_PURPLE, "N": _FILL_TOT_PURPLE,
    "O": _FILL_TOT_GOLD, "P": _FILL_TOT_GOLD,
    "Q": _FILL_TOT_GOLD, "R": _FILL_TOT_GOLD,
}

# Columns with right thin border separator (visual column grouping)
_SEP_COLS = {"D", "J", "N", "T", "X"}


# ===========================================================================
#  Public API
# ===========================================================================

def build_dsr(
    spine: Spine,
    report_date: date,
    rc_data: pd.DataFrame | None = None,
    hs_data: pd.DataFrame | None = None,
    nb_data: pd.DataFrame | None = None,
    quotes_data: pd.DataFrame | None = None,
    premium_data: pd.DataFrame | None = None,
    screenshot_data: pd.DataFrame | None = None,
    output_folder: str = "reports",
) -> Path:
    agents = pd.DataFrame(spine.all_agents())
    merged = agents.copy()

    if rc_data is not None and len(rc_data) > 0:
        rc_agg = rc_data.groupby("Agent").agg(
            Calls=("Calls", "sum"), Inbound=("Inbound", "sum"),
            Outbound=("Outbound", "sum"), TalkTimeSeconds=("TalkTimeSeconds", "sum"),
        ).reset_index()
        merged = merged.merge(rc_agg, left_on="agent", right_on="Agent", how="left", suffixes=("", "_rc"))

    if hs_data is not None and len(hs_data) > 0:
        hs_agg = hs_data.groupby("Agent").agg(
            Texts=("Texts", "sum"), OutTexts=("OutTexts", "sum"),
            OptIns=("OptIns", "sum"), OptOuts=("OptOuts", "sum"),
        ).reset_index()
        merged = merged.merge(hs_agg, left_on="agent", right_on="Agent", how="left", suffixes=("", "_hs"))

    if quotes_data is not None and len(quotes_data) > 0:
        q_agg = quotes_data.groupby("Agent").agg(Quotes=("QuoteCount", "sum")).reset_index()
        merged = merged.merge(q_agg, left_on="agent", right_on="Agent", how="left", suffixes=("", "_q"))

    if nb_data is not None and len(nb_data) > 0:
        nb_agg = nb_data.groupby("Agent").agg(
            NB=("NBCount", "sum"), Items=("Items", "sum"),
            WrittenPremium=("WrittenPremium", "sum"),
        ).reset_index()
        merged = merged.merge(nb_agg, left_on="agent", right_on="Agent", how="left", suffixes=("", "_nb"))

    if screenshot_data is not None and len(screenshot_data) > 0:
        ss_cols = ["Agent"]
        for col in ["dismissed_todos", "past_due_todos", "contact", "quoted", "hot", "xsale"]:
            if col in screenshot_data.columns:
                ss_cols.append(col)
        ss_agg = screenshot_data[ss_cols].groupby("Agent").first().reset_index()
        merged = merged.merge(ss_agg, left_on="agent", right_on="Agent", how="left", suffixes=("", "_ss"))

    merged = merged.fillna(0)

    rows = []
    month_start = report_date.replace(day=1)
    for _, row in merged.sort_values(["team", "agent"]).iterrows():
        talk_seconds = int(row.get("TalkTimeSeconds", 0))
        rows.append({
            "Date": report_date,
            "Team": row.get("team", ""),
            "Agent": row.get("agent", ""),
            "Office": row.get("office", ""),
            "Full Name": row.get("full_name", ""),
            "Calls": int(row.get("Calls", 0)),
            "Inbound": int(row.get("Inbound", 0)),
            "Outbound": int(row.get("Outbound", 0)),
            "Talk Time": _seconds_to_time(talk_seconds),
            "Texts": int(row.get("Texts", 0)),
            "OutTexts": int(row.get("OutTexts", 0)),
            "Opt-Ins": int(row.get("OptIns", 0)),
            "Opt-Outs": int(row.get("OptOuts", 0)),
            "Quotes": int(row.get("Quotes", 0)),
            "NB": int(row.get("NB", 0)),
            "Total Premium": round(float(row.get("WrittenPremium", 0)), 2),
            "Items": int(row.get("Items", 0)),
            "MonthStart": month_start,
            "Items MTD": int(row.get("ItemsMTD", 0)),
            "Contact": int(row.get("contact", 0)),
            "Quoted": int(row.get("quoted", 0)),
            "Hot": int(row.get("hot", 0)),
            "XDate": int(row.get("xsale", 0)),
            "Dismissed To-Do's": int(row.get("dismissed_todos", 0)),
            "Past Due To-Do's": int(row.get("past_due_todos", 0)),
        })

    output_path = Path(output_folder) / f"DSR_{report_date.isoformat()}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "DSR"

    _write_dsr_sheet(ws, rows, report_date)

    wb.save(str(output_path))
    print(f"[dsr_builder] Report saved: {output_path}")
    return output_path


# ===========================================================================
#  Sheet writer
# ===========================================================================

def _write_dsr_sheet(ws, rows: list[dict], report_date: date):
    last_data_row = len(rows) + 3  # header=3, data starts=4

    # --- Sheet-level settings ---
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_view.zoomScaleNormal = 90
    ws.sheet_properties.tabColor = "FFC000"

    # --- Column widths ---
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # --- Hidden columns ---
    for col_letter in HIDDEN_COLS:
        ws.column_dimensions[col_letter].hidden = True

    # --- Default row height ---
    ws.sheet_format.defaultRowHeight = 18.0

    # =====================================================
    #  ROW 1: Title
    # =====================================================
    ws.merge_cells("B1:N1")
    ws["A1"].font = _FNT_A1  # red bold (matches original)
    ws["B1"] = "Daily Standup Report"
    ws["B1"].font = _FNT_TITLE
    ws["B1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 33.75

    ws["T1"] = "Agency Auto Items MTD"
    ws["T1"].font = _FNT_T1
    ws["T1"].fill = _FILL_T1_LABEL
    ws["T1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # =====================================================
    #  ROW 2: Subtotals (above the table, uses SUBTOTAL)
    # =====================================================
    ws.row_dimensions[2].height = 16.2

    # Subtotal formulas for each numeric column
    subtotal_cols = {
        "G": ("General", _FILL_TOT_GREEN, _FNT_TOT_WHITE),
        "H": ("General", _FILL_TOT_GREEN, _FNT_TOT_WHITE),
        "I": ("General", _FILL_TOT_GREEN, _FNT_TOT_WHITE),
        "J": ("d\\.hh:mm:ss", _FILL_TOT_GREEN, _FNT_TOT_WHITE),
        "K": ("General", _FILL_TOT_PURPLE, _FNT_TOT_WHITE),
        "L": ("General", _FILL_TOT_PURPLE, _FNT_TOT_WHITE),
        "M": ("General", _FILL_TOT_PURPLE, _FNT_TOT_WHITE),
        "N": ("General", _FILL_TOT_PURPLE, _FNT_TOT_WHITE),
        "O": ("General", _FILL_TOT_GOLD, _FNT_TOT_DARK),
        "P": ("General", _FILL_TOT_GOLD, _FNT_TOT_DARK),
        "Q": ('"$"#,##0.00', _FILL_TOT_GOLD, _FNT_TOT_DARK),
        "R": ("0", _FILL_TOT_GOLD, _FNT_TOT_DARK),
    }

    for col_letter, (num_fmt, fill, font) in subtotal_cols.items():
        cell = ws[f"{col_letter}2"]
        cell.value = f"=SUBTOTAL(109,{col_letter}4:{col_letter}{last_data_row})"
        cell.number_format = num_fmt
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # Borders on row 2
    # G2: left+top+bottom medium
    ws["G2"].border = Border(left=_MED, top=_MED, bottom=_MED)
    # H2-N2: top+bottom medium
    for c in ["H", "I", "J", "K", "L", "M", "N"]:
        ws[f"{c}2"].border = Border(top=_MED, bottom=_MED)
    # O2-R2: top+bottom medium (bottom=medium)
    for c in ["O", "P", "Q", "R"]:
        ws[f"{c}2"].border = Border(top=_MED, bottom=_MED)

    # T2: static value with all-sides medium border, yellow fill
    ws["T2"] = f"=SUBTOTAL(109,T4:T{last_data_row})"
    ws["T2"].number_format = "0"
    ws["T2"].font = _FNT_TOT_DARK
    ws["T2"].fill = _FILL_TOT_YELLOW
    ws["T2"].alignment = Alignment(horizontal="center")
    ws["T2"].border = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

    # Y2:Z2 "eAgent only" merged label
    ws.merge_cells("Y2:Z2")
    ws["Y2"] = "eAgent only"
    ws["Y2"].font = _FNT_EAGENT
    ws["Y2"].fill = _FILL_EAGENT
    ws["Y2"].alignment = Alignment(horizontal="center", wrap_text=True)

    # =====================================================
    #  ROW 3: Headers
    # =====================================================
    ws.row_dimensions[3].height = 16.2

    for col_idx, col_name in enumerate(DSR_COLUMNS, start=2):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=3, column=col_idx, value=col_name)

        # Font: default 11pt, purple headers 12pt bold white, Y-Z 10pt
        font = _HDR_FONT.get(col_letter, _FNT_HDR_DEFAULT)
        cell.font = font

        # Fill
        fill = _HDR_FILL.get(col_letter)
        if fill:
            cell.fill = fill

        # Alignment: all wrap, K-N left-aligned, rest default
        if col_letter in ("K", "L", "M", "N"):
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
        else:
            cell.alignment = Alignment(wrap_text=True)

    # Header borders
    # B3: left+top+bottom medium
    ws["B3"].border = Border(left=_MED, top=_MED, bottom=_MED)
    # C3: top+bottom medium
    ws["C3"].border = Border(top=_MED, bottom=_MED)
    # D3: top+bottom medium, right thin
    ws["D3"].border = Border(top=_MED, bottom=_MED, right=_THIN)
    # E3-F3: top+bottom medium
    for c in ["E", "F"]:
        ws[f"{c}3"].border = Border(top=_MED, bottom=_MED)
    # G3-I3: no explicit border (table style handles it)
    # J3: right thin, top medium
    ws["J3"].border = Border(right=_THIN, top=_MED)
    # K3-M3: no border
    # N3: right thin, top medium
    ws["N3"].border = Border(right=_THIN, top=_MED)
    # O3-R3: bottom medium
    for c in ["O", "P", "Q", "R"]:
        ws[f"{c}3"].border = Border(bottom=_MED)
    # S3: top+bottom medium
    ws["S3"].border = Border(top=_MED, bottom=_MED)
    # T3: right thin, top+bottom medium
    ws["T3"].border = Border(right=_THIN, top=_MED, bottom=_MED)
    # U3-W3: top+bottom medium
    for c in ["U", "V", "W"]:
        ws[f"{c}3"].border = Border(top=_MED, bottom=_MED)
    # X3: right thin, top+bottom medium
    ws["X3"].border = Border(right=_THIN, top=_MED, bottom=_MED)

    # =====================================================
    #  DATA ROWS (row 4+)
    # =====================================================
    for row_idx, row_data in enumerate(rows, start=4):
        ws.row_dimensions[row_idx].height = 15.6

        for col_idx, col_name in enumerate(DSR_COLUMNS, start=2):
            col_letter = get_column_letter(col_idx)
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Font: Q (Premium) and Y-Z (To-Do's) = 11pt, rest = 12pt
            if col_name in ("Total Premium", "Dismissed To-Do's", "Past Due To-Do's"):
                cell.font = _FNT_DATA_SM
            elif col_name == "Date":
                cell.font = _FNT_DATA_SM
            else:
                cell.font = _FNT_DATA

            # Alignment: all centered
            cell.alignment = Alignment(horizontal="center")

            # Number formats
            if col_name == "Date":
                cell.number_format = 'm/d\\, ddd'
            elif col_name == "MonthStart":
                cell.number_format = 'mm-dd-yy'
            elif col_name == "Talk Time":
                cell.number_format = 'hh:mm'
            elif col_name == "Total Premium":
                cell.number_format = '"$"#,##0.00'

            # Border separators (thin right) at group boundaries
            if col_letter in _SEP_COLS:
                cell.border = Border(right=_THIN)

        # B column: left thin border
        ws.cell(row=row_idx, column=2).border = Border(left=_THIN)

    # =====================================================
    #  EXCEL TABLE (for auto-filter + row banding + slicer support)
    # =====================================================
    table_ref = f"B3:Z{last_data_row}"
    tab = Table(displayName="DSR", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium18",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # =====================================================
    #  CONDITIONAL FORMATTING
    # =====================================================

    # 1. Team column: "Sales" = dark green, "CSR" = dark purple
    team_range = f"C4:C{last_data_row}"
    ws.conditional_formatting.add(
        team_range,
        CellIsRule(
            operator="equal",
            formula=['"Sales"'],
            font=Font(name="Aptos Narrow", size=12, bold=True, color="3B7D1E"),
        ),
    )
    ws.conditional_formatting.add(
        team_range,
        CellIsRule(
            operator="equal",
            formula=['"CSR"'],
            font=Font(name="Aptos Narrow", size=12, bold=True, color="781F6E"),
        ),
    )

    # 2. Zeros in numeric range = gray (fade out)
    data_range = f"G4:Z{last_data_row}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator="equal",
            formula=["0"],
            font=Font(color="A6A6A6"),
        ),
    )

    # 3. Quotes (O) and Items (R): zeros = salmon pink
    for col in ["O", "R"]:
        ws.conditional_formatting.add(
            f"{col}4:{col}{last_data_row}",
            CellIsRule(
                operator="equal",
                formula=["0"],
                font=Font(color="F69C9C"),
            ),
        )


def _seconds_to_time(seconds: int) -> time:
    """Convert total seconds to a time object (capped at 23:59:59)."""
    if seconds <= 0:
        return time(0, 0, 0)
    hours = min(seconds // 3600, 23)
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    return time(hours, minutes, secs)
