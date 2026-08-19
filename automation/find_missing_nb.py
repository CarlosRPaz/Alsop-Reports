"""
Compare NB data from BOTH Excel files against DB for every May day.
Find any missing days/policies.
"""
import openpyxl
import requests
from collections import defaultdict, Counter
from datetime import datetime

QUOTES_PATH = r"c:\Users\scag3s29\Documents\Claude Scope\Quotes Detail Report__New Business Detail.xlsx"
DSR_PATH = r"c:\Users\scag3s29\Documents\Claude Scope\Daily Standup Report.xlsx"
URL = "https://xejmpdfqaghamemjrhxa.supabase.co"
KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inhlam1wZGZxYWdoYW1lbWpyaHhhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY0NTkxMTUsImV4cCI6MjA5MjAzNTExNX0.0m_8BHyk-2dVZUjCme-yDXwrpswhpBi8gFZVTdIyWOc"
HEADERS = {"apikey": KEY, "Authorization": f"Bearer {KEY}"}

def parse_date(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None

# ── Parse Quotes Detail Report NB sheet ──
print("=== QUOTES DETAIL REPORT (New Business sheet) ===")
wb1 = openpyxl.load_workbook(QUOTES_PATH, read_only=True, data_only=True)
ws1 = wb1["New Business"]
q_by_day = defaultdict(lambda: {"policies": 0, "items": 0})
for row in ws1.iter_rows(min_row=6, values_only=True):
    d = parse_date(row[7])
    if d and d.year == 2026 and d.month == 5:
        day_str = d.strftime("%Y-%m-%d")
        q_by_day[day_str]["policies"] += 1
        q_by_day[day_str]["items"] += row[14] or 0
wb1.close()

for d in sorted(q_by_day):
    print(f"  {d}: {q_by_day[d]['policies']} policies, {q_by_day[d]['items']} items")
print(f"  TOTAL: {sum(v['policies'] for v in q_by_day.values())} policies, {sum(v['items'] for v in q_by_day.values())} items")

# ── Parse DSR NB sheet ──
print("\n=== DSR FILE (NB sheet) ===")
wb2 = openpyxl.load_workbook(DSR_PATH, data_only=True, read_only=False)
ws2 = wb2["NB"]
dsr_by_day = defaultdict(lambda: {"policies": 0, "items": 0})
for row in ws2.iter_rows(min_row=2, values_only=True):
    d = parse_date(row[7])
    if d and d.year == 2026 and d.month == 5:
        day_str = d.strftime("%Y-%m-%d")
        dsr_by_day[day_str]["policies"] += 1
        dsr_by_day[day_str]["items"] += row[14] or 0
wb2.close()

for d in sorted(dsr_by_day):
    print(f"  {d}: {dsr_by_day[d]['policies']} policies, {dsr_by_day[d]['items']} items")
print(f"  TOTAL: {sum(v['policies'] for v in dsr_by_day.values())} policies, {sum(v['items'] for v in dsr_by_day.values())} items")

# ── Get DB ──
print("\n=== DATABASE (daily_metrics) ===")
res = requests.get(
    f"{URL}/rest/v1/daily_metrics?select=report_date,nb_count,items"
    f"&report_date=gte.2026-05-01&report_date=lte.2026-05-14",
    headers=HEADERS
)
db_rows = res.json()
db_by_day = defaultdict(lambda: {"nb_count": 0, "items": 0})
for r in db_rows:
    db_by_day[r["report_date"]]["nb_count"] += r["nb_count"] or 0
    db_by_day[r["report_date"]]["items"] += r["items"] or 0

for d in sorted(db_by_day):
    if db_by_day[d]["nb_count"] > 0 or db_by_day[d]["items"] > 0:
        print(f"  {d}: {db_by_day[d]['nb_count']} nb_count, {db_by_day[d]['items']} items")
print(f"  TOTAL: {sum(v['nb_count'] for v in db_by_day.values())} nb, {sum(v['items'] for v in db_by_day.values())} items")

# ── Full comparison ──
all_dates = sorted(set(
    list(q_by_day.keys()) + list(dsr_by_day.keys()) + 
    [d for d in db_by_day if db_by_day[d]["nb_count"] > 0]
))

print(f"\n\n{'='*85}")
print(f"{'Date':<12} {'QuotesFile':>11} {'DSR File':>10} {'DB nb':>7} {'DB items':>9} {'DSR items':>10} {'Q items':>8}")
print(f"{'='*85}")

for d in all_dates:
    qp = q_by_day[d]["policies"]
    qi = q_by_day[d]["items"]
    dp = dsr_by_day[d]["policies"]
    di = dsr_by_day[d]["items"]
    dbn = db_by_day[d]["nb_count"]
    dbi = db_by_day[d]["items"]
    
    # Flag mismatches
    flags = []
    if dp > 0 and dbn != dp:
        flags.append("DB!=DSR")
    if dp > 0 and di != dbi:
        flags.append("items!=DSR")
    flag = " *** " + ",".join(flags) if flags else ""
    
    print(f"  {d:<10} {qp:>11} {dp:>10} {dbn:>7} {dbi:>9} {di:>10} {qi:>8}{flag}")

# Show days of month with NO NB at all
print(f"\n=== BUSINESS DAYS WITH NO NB IN ANY SOURCE ===")
from datetime import date, timedelta
d = date(2026, 5, 1)
end = date(2026, 5, 14)
while d <= end:
    ds = d.strftime("%Y-%m-%d")
    dow = d.strftime("%A")
    if d.weekday() < 5:  # business day
        has_q = q_by_day[ds]["policies"] > 0
        has_dsr = dsr_by_day[ds]["policies"] > 0
        has_db = db_by_day[ds]["nb_count"] > 0
        if not has_q and not has_dsr and not has_db:
            print(f"  {ds} ({dow}): NO NB DATA IN ANY SOURCE")
        elif not has_db:
            print(f"  {ds} ({dow}): Missing from DB! (QuotesFile={q_by_day[ds]['policies']}, DSR={dsr_by_day[ds]['policies']})")
    d += timedelta(days=1)
